import os
import re
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Annotated

import duckdb
from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from pydantic import Field

mcp = FastMCP("mcp-server-xlsx")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load(file: str) -> Workbook:
    p = Path(file)
    if not p.exists():
        raise ValueError(f"File not found: {file}")
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"Not an .xlsx file: {file}")
    return load_workbook(p, data_only=False)


def _resolve_sheet(wb: Workbook, sheet: str | None):
    if sheet is None:
        return wb.worksheets[0]
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet!r}. Available: {wb.sheetnames}")
    return wb[sheet]


def _save(wb: Workbook, file: str) -> None:
    p = Path(file)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=p.parent)
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, p)
    except BaseException:
        os.unlink(tmp)
        raise


def _parse_cell(cell: str) -> tuple[int, int]:
    cell = cell.replace("$", "")
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell)
    if not match:
        raise ValueError(f"Invalid cell reference: {cell!r}")
    col = column_index_from_string(match.group(1))
    row = int(match.group(2))
    return row, col


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_str.replace("$", ""))
    except Exception:
        raise ValueError(f"Invalid range: {range_str!r}")
    return min_col, min_row, max_col, max_row


def _coerce_value(value):
    if value is None or not isinstance(value, str):
        return value
    if value.startswith("="):
        return value
    try:
        if "." in value or "e" in value.lower():
            return float(value)
        return int(value)
    except ValueError:
        return value


# ---------------------------------------------------------------------------
# Workbook Operations
# ---------------------------------------------------------------------------

@mcp.tool()
def list_workbooks(
    directory: Annotated[str, Field(description="Absolute or relative path to the directory to scan")],
) -> list[str]:
    """List all .xlsx files in a directory (non-recursive).

    Returns the full path of each .xlsx file found, sorted alphabetically.
    """
    d = Path(directory)
    if not d.is_dir():
        raise ValueError(f"Not a directory: {directory}")
    return sorted(str(f) for f in d.iterdir() if f.suffix.lower() == ".xlsx" and f.is_file())


@mcp.tool()
def create_workbook(
    file: Annotated[str, Field(description="Path where the new .xlsx file will be created. Must not already exist.")],
    sheet_name: Annotated[str | None, Field(description="Name for the initial sheet. Defaults to 'Sheet' if omitted.")] = None,
) -> str:
    """Create a new empty .xlsx workbook at the given path.

    The file must not already exist. Returns the absolute path of the
    created file. The workbook is created with a single empty sheet.
    """
    p = Path(file)
    if p.exists():
        raise ValueError(f"File already exists: {file}")
    wb = Workbook()
    if sheet_name:
        wb.active.title = sheet_name
    _save(wb, file)
    return str(p.resolve())


@mcp.tool()
def copy_workbook(
    source: Annotated[str, Field(description="Path to the existing .xlsx file to copy")],
    destination: Annotated[str, Field(description="Path for the new copy. Must not already exist.")],
) -> str:
    """Copy an existing .xlsx file to a new location.

    Performs a full file copy preserving all data, formatting, and metadata.
    The destination must not already exist. Returns the absolute path of
    the new file.
    """
    src = Path(source)
    if not src.exists():
        raise ValueError(f"Source not found: {source}")
    dst = Path(destination)
    if dst.exists():
        raise ValueError(f"Destination already exists: {destination}")
    shutil.copy2(src, dst)
    return str(dst.resolve())


# ---------------------------------------------------------------------------
# Sheet Operations
# ---------------------------------------------------------------------------

@mcp.tool()
def list_sheets(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
) -> list[str]:
    """List all sheet names in a workbook, in workbook order.

    Returns a list of sheet name strings, e.g. ["Sheet1", "Data", "Summary"].
    """
    wb = _load(file)
    return wb.sheetnames


@mcp.tool()
def add_sheet(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    name: Annotated[str | None, Field(description="Name for the new sheet. Auto-generated if omitted (e.g. 'Sheet1').")] = None,
    position: Annotated[int | None, Field(description="1-based position to insert the sheet. Appended at the end if omitted.")] = None,
) -> str:
    """Add a new sheet to the workbook.

    Returns the name of the newly created sheet (which may differ from the
    requested name if it conflicted with an existing sheet).
    """
    wb = _load(file)
    ws = wb.create_sheet(title=name, index=position if position is None else position - 1)
    _save(wb, file)
    return ws.title


@mcp.tool()
def rename_sheet(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    old_name: Annotated[str, Field(description="Current name of the sheet to rename")],
    new_name: Annotated[str, Field(description="New name for the sheet")],
) -> str:
    """Rename an existing sheet in the workbook.

    Returns the new sheet name on success.
    """
    wb = _load(file)
    if old_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {old_name!r}")
    wb[old_name].title = new_name
    _save(wb, file)
    return new_name


@mcp.tool()
def delete_sheet(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    name: Annotated[str, Field(description="Name of the sheet to delete")],
) -> str:
    """Delete a sheet by name from the workbook.

    All data in the sheet is permanently removed. The workbook must have
    at least one remaining sheet after deletion.
    """
    wb = _load(file)
    if name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {name!r}")
    del wb[name]
    _save(wb, file)
    return f"Deleted sheet {name!r}"


@mcp.tool()
def copy_sheet(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    source_name: Annotated[str, Field(description="Name of the existing sheet to duplicate")],
    new_name: Annotated[str | None, Field(description="Name for the copy. Auto-generated (e.g. 'Sheet Copy') if omitted.")] = None,
    position: Annotated[int | None, Field(description="1-based position for the copied sheet. Placed at the end if omitted.")] = None,
) -> str:
    """Duplicate a sheet within the same workbook.

    Copies all cell values and formatting. Returns the name of the new sheet.
    """
    wb = _load(file)
    if source_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {source_name!r}")
    copied = wb.copy_worksheet(wb[source_name])
    if new_name:
        copied.title = new_name
    if position is not None:
        wb.move_sheet(copied, offset=position - 1 - wb.sheetnames.index(copied.title))
    _save(wb, file)
    return copied.title


# ---------------------------------------------------------------------------
# Reading Data
# ---------------------------------------------------------------------------

@mcp.tool()
def read_sheet(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
    start_row: Annotated[int | None, Field(description="First row to include (1-based). Defaults to the first used row.")] = None,
    end_row: Annotated[int | None, Field(description="Last row to include (1-based). Defaults to the last used row.")] = None,
    start_column: Annotated[int | None, Field(description="First column to include (1-based, e.g. 1 = A). Defaults to the first used column.")] = None,
    end_column: Annotated[int | None, Field(description="Last column to include (1-based). Defaults to the last used column.")] = None,
) -> list[list]:
    """Read an entire sheet (or a bounded sub-region) as a list of rows.

    Each row is a list of cell values. Formula cells return the formula
    string (e.g. '=SUM(A1:A5)'), not the cached result. Empty cells
    appear as null. Use the optional row/column bounds to limit output
    for large sheets.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    rows = ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
        values_only=True,
    )
    return [list(r) for r in rows]


@mcp.tool()
def read_cell(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    cell: Annotated[str, Field(description="Cell reference in Excel notation, e.g. 'B3' or '$B$3'")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
):
    """Read the value of a single cell.

    Returns the cell's value in its stored type: numbers as int/float,
    text as string, formulas as their formula string (e.g. '=A1+B1'),
    and empty cells as null.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    row, col = _parse_cell(cell)
    return ws.cell(row=row, column=col).value


@mcp.tool()
def read_range(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    range_str: Annotated[str, Field(description="Cell range in Excel notation, e.g. 'A1:D10' or '$A$1:$D$10'")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> list[list]:
    """Read a rectangular range of cells as a list of rows.

    Returns a 2D array where each inner list is one row of values.
    Formula cells return the formula string, empty cells return null.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    min_col, min_row, max_col, max_row = _parse_range(range_str)
    rows = ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )
    return [list(r) for r in rows]


@mcp.tool()
def get_sheet_dimensions(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> dict:
    """Get the dimensions of the used range in a sheet.

    Returns {"rows": N, "columns": M} where N is the number of the last
    used row and M is the number of the last used column. Both are 0 for
    an empty sheet.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    return {
        "rows": ws.max_row or 0,
        "columns": ws.max_column or 0,
    }


# ---------------------------------------------------------------------------
# Writing Data
# ---------------------------------------------------------------------------

@mcp.tool()
def write_cell(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    cell: Annotated[str, Field(description="Target cell in Excel notation, e.g. 'B3'")],
    value: Annotated[object, Field(description="Value to write. Numeric strings are coerced to numbers, strings starting with '=' are stored as formulas, everything else is stored as text.")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Write a single value to a cell.

    Overwrites any existing value. The value is type-coerced: numeric
    strings become numbers, '=' prefix means formula, all else is text.
    Existing cell formatting is preserved.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    row, col = _parse_cell(cell)
    ws.cell(row=row, column=col, value=_coerce_value(value))
    _save(wb, file)
    return f"Wrote to {cell}"


@mcp.tool()
def write_range(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    start_cell: Annotated[str, Field(description="Top-left cell where writing begins, e.g. 'B2'")],
    data: Annotated[list[list], Field(description="2D array of values (list of rows), e.g. [[1, 2, 3], ['a', 'b', 'c']]. Numeric strings are coerced to numbers, '=' prefix is treated as a formula.")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Write a 2D array of values into a rectangular region.

    Writing starts at start_cell and expands right and down to fit the
    data. Each value is type-coerced: numeric strings become numbers,
    strings starting with '=' are stored as formulas, everything else
    is text. Prefer this over multiple write_cell calls for efficiency.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    row_off, col_off = _parse_cell(start_cell)
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            ws.cell(row=row_off + r_idx, column=col_off + c_idx, value=_coerce_value(val))
    _save(wb, file)
    return f"Wrote {len(data)} rows starting at {start_cell}"


@mcp.tool()
def append_rows(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    data: Annotated[list[list], Field(description="List of rows to append, e.g. [['Alice', 30], ['Bob', 25]]. Each inner list is one row.")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Append one or more rows after the last used row in the sheet.

    Values are type-coerced (numeric strings to numbers, '=' to formulas).
    This is the most efficient way to add data to the end of a sheet.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    for row in data:
        ws.append([_coerce_value(v) for v in row])
    _save(wb, file)
    return f"Appended {len(data)} rows"


@mcp.tool()
def insert_rows(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    row: Annotated[int, Field(description="1-based row index where new rows will be inserted. Existing rows at and below this index shift down.")],
    count: Annotated[int, Field(description="Number of rows to insert. If data is provided and longer, enough rows are inserted to fit the data.")] = 1,
    data: Annotated[list[list] | None, Field(description="Optional 2D array of values to fill the inserted rows, e.g. [['a', 'b'], ['c', 'd']]. Leave empty for blank rows.")] = None,
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Insert rows at a given position, shifting existing rows down.

    If data is provided, the inserted rows are filled with those values
    (type-coerced). Otherwise the rows are left blank.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    actual_count = max(count, len(data) if data else 0)
    ws.insert_rows(row, amount=actual_count)
    if data:
        for r_idx, row_data in enumerate(data):
            for c_idx, val in enumerate(row_data):
                ws.cell(row=row + r_idx, column=c_idx + 1, value=_coerce_value(val))
    _save(wb, file)
    return f"Inserted {actual_count} rows at row {row}"


@mcp.tool()
def delete_rows(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    row: Annotated[int, Field(description="1-based index of the first row to delete")],
    count: Annotated[int, Field(description="Number of consecutive rows to delete starting from row")] = 1,
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Delete one or more rows, shifting remaining rows up.

    All data in the deleted rows is permanently removed. Cell references
    in formulas on other sheets are not automatically updated.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    ws.delete_rows(row, amount=count)
    _save(wb, file)
    return f"Deleted {count} rows starting at row {row}"


@mcp.tool()
def clear_range(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    range_str: Annotated[str, Field(description="Range to clear in Excel notation, e.g. 'A1:D10'. Cell formatting is preserved, only values are removed.")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Clear all cell values in a range without removing rows or columns.

    Sets every cell in the range to null. Row/column structure and cell
    formatting are preserved — only the values are erased.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    min_col, min_row, max_col, max_row = _parse_range(range_str)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).value = None
    _save(wb, file)
    return f"Cleared range {range_str}"


@mcp.tool()
def copy_range(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    source_range: Annotated[str, Field(description="Range to copy from in Excel notation, e.g. 'A1:C5'")],
    dest_start: Annotated[str, Field(description="Top-left cell of the destination, e.g. 'E1'. The copied block expands right and down from here.")],
    sheet: Annotated[str | None, Field(description="Source sheet name. Defaults to the first sheet if omitted.")] = None,
    dest_sheet: Annotated[str | None, Field(description="Destination sheet name. Defaults to the same sheet as the source if omitted.")] = None,
) -> str:
    """Copy a rectangular block of cells to another location.

    Copies raw values only (not formatting). The destination can be on
    the same sheet or a different sheet in the same workbook. Existing
    values at the destination are overwritten.
    """
    wb = _load(file)
    src_ws = _resolve_sheet(wb, sheet)
    dst_ws = _resolve_sheet(wb, dest_sheet) if dest_sheet else src_ws

    min_col, min_row, max_col, max_row = _parse_range(source_range)
    dest_row, dest_col = _parse_cell(dest_start)

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            val = src_ws.cell(row=r, column=c).value
            dst_ws.cell(
                row=dest_row + (r - min_row),
                column=dest_col + (c - min_col),
                value=val,
            )
    _save(wb, file)
    return f"Copied {source_range} to {dest_start}"


# ---------------------------------------------------------------------------
# Column Operations
# ---------------------------------------------------------------------------

@mcp.tool()
def insert_columns(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    column: Annotated[int, Field(description="1-based column index where new columns will be inserted (e.g. 1 = A, 2 = B). Existing columns at and to the right shift right.")],
    count: Annotated[int, Field(description="Number of blank columns to insert")] = 1,
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Insert one or more blank columns, shifting existing columns right.

    The inserted columns are empty. Existing data to the right of the
    insertion point is shifted.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    ws.insert_cols(column, amount=count)
    _save(wb, file)
    return f"Inserted {count} columns at column {column}"


@mcp.tool()
def delete_columns(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    column: Annotated[int, Field(description="1-based index of the first column to delete (e.g. 1 = A, 2 = B)")],
    count: Annotated[int, Field(description="Number of consecutive columns to delete starting from column")] = 1,
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> str:
    """Delete one or more columns, shifting remaining columns left.

    All data in the deleted columns is permanently removed.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    ws.delete_cols(column, amount=count)
    _save(wb, file)
    return f"Deleted {count} columns at column {column}"


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

@mcp.tool()
def search_sheet(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    pattern: Annotated[str, Field(description="Regular expression pattern to search for. Matched against the string representation of each cell value. Use simple strings for literal search, or regex syntax for pattern matching (e.g. '\\d{3}-\\d{4}' for phone-number fragments).")],
    sheet: Annotated[str | None, Field(description="Sheet name. Defaults to the first sheet if omitted.")] = None,
) -> list[dict]:
    """Search all cells in a sheet for values matching a regex pattern.

    Returns a list of matches, each with the cell reference and value,
    e.g. [{"cell": "B3", "value": "hello"}, ...]. Searches the string
    representation of every non-empty cell. Returns an empty list if
    no matches are found.
    """
    wb = _load(file)
    ws = _resolve_sheet(wb, sheet)
    regex = re.compile(pattern)
    results = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            val = cell.value
            if val is None:
                continue
            if regex.search(str(val)):
                results.append({
                    "cell": f"{get_column_letter(cell.column)}{cell.row}",
                    "value": val,
                })
    return results


# ---------------------------------------------------------------------------
# Table Mode — SQL Helpers
# ---------------------------------------------------------------------------

def _sheet_to_records(ws, header_row: int = 1) -> tuple[list[str], list[tuple]]:
    """Extract headers and data rows from a worksheet.

    Returns (headers, rows). Skips fully-empty data rows.
    """
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0

    if max_col == 0:
        return [], []

    headers = []
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            break
        headers.append(str(val))

    if not headers:
        return [], []

    num_cols = len(headers)
    rows = []
    for row_idx in range(header_row + 1, max_row + 1):
        row = tuple(ws.cell(row=row_idx, column=c).value for c in range(1, num_cols + 1))
        if all(v is None for v in row):
            continue
        rows.append(row)

    return headers, rows


def _dedup_headers(headers: list[str]) -> list[str]:
    """Deduplicate column names by appending _N suffixes."""
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


def _infer_duckdb_type(values) -> str:
    """Infer a DuckDB column type from a list of Python values."""
    types_seen: set[str] = set()
    for v in values:
        if v is None:
            continue
        if isinstance(v, bool):
            types_seen.add("BOOLEAN")
        elif isinstance(v, int):
            types_seen.add("BIGINT")
        elif isinstance(v, float):
            types_seen.add("DOUBLE")
        elif isinstance(v, datetime):
            types_seen.add("TIMESTAMP")
        elif isinstance(v, date):
            types_seen.add("DATE")
        else:
            types_seen.add("VARCHAR")

    if not types_seen:
        return "VARCHAR"
    if len(types_seen) == 1:
        return types_seen.pop()
    if types_seen <= {"BIGINT", "DOUBLE"}:
        return "DOUBLE"
    return "VARCHAR"


def _load_sheets_to_duckdb(wb: Workbook, header_row: int = 1) -> duckdb.DuckDBPyConnection:
    """Load all sheets into an in-memory DuckDB database.

    Each sheet becomes a table named after its sheet title.
    """
    conn = duckdb.connect()

    for ws in wb.worksheets:
        headers, rows = _sheet_to_records(ws, header_row)
        if not headers:
            continue

        headers = _dedup_headers(headers)
        num_cols = len(headers)

        if rows:
            columns_data = list(zip(*rows))
            types = [_infer_duckdb_type(col) for col in columns_data]
        else:
            types = ["VARCHAR"] * num_cols

        col_defs = ", ".join(f'"{h}" {t}' for h, t in zip(headers, types))
        conn.execute(f'CREATE TABLE "{ws.title}" ({col_defs})')

        if rows:
            placeholders = ", ".join(["?"] * num_cols)
            conn.executemany(f'INSERT INTO "{ws.title}" VALUES ({placeholders})', rows)

    return conn


def _extract_target_table(sql: str) -> str:
    """Extract the target table name from a mutating SQL statement."""
    s = sql.strip()
    patterns = [
        r"INSERT\s+INTO\s+",
        r"UPDATE\s+",
        r"DELETE\s+FROM\s+",
    ]
    for pat in patterns:
        m = re.match(pat + r'"([^"]+)"', s, re.IGNORECASE)
        if m:
            return m.group(1)
        m = re.match(pat + r"(\w+)", s, re.IGNORECASE)
        if m:
            return m.group(1)
    raise ValueError(
        "Could not determine target table. "
        "SQL must start with INSERT INTO, UPDATE, or DELETE FROM."
    )


def _infer_describe_type(values) -> str:
    """Infer a human-readable column type for describe_table."""
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "unknown"
    if all(isinstance(v, bool) for v in non_null):
        return "boolean"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_null):
        return "integer"
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_null):
        return "number"
    if all(isinstance(v, (datetime, date)) for v in non_null):
        return "date"
    return "text"


# ---------------------------------------------------------------------------
# Table Mode — SQL Tools
# ---------------------------------------------------------------------------

@mcp.tool()
def describe_table(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    sheet: Annotated[str | None, Field(description="Sheet name to describe. If omitted, describes all sheets in the workbook.")] = None,
    header_row: Annotated[int, Field(description="1-based row number containing column headers. Defaults to 1.")] = 1,
) -> list[dict] | dict:
    """Inspect the structure of a sheet treated as a database table.

    Returns column names, inferred data types (text, integer, number,
    boolean, date), total row count, and sample values from the first 3
    data rows. Use this before writing SQL queries to understand the
    available columns and their types.

    When sheet is omitted, returns a list of descriptions for all sheets.
    """
    wb = _load(file)
    targets = [_resolve_sheet(wb, sheet)] if sheet else wb.worksheets

    results = []
    for ws in targets:
        headers, rows = _sheet_to_records(ws, header_row)

        if not headers:
            results.append({"sheet": ws.title, "columns": [], "row_count": 0, "sample": []})
            continue

        columns = []
        for col_idx, header in enumerate(headers):
            col_values = [row[col_idx] for row in rows]
            columns.append({"name": header, "type": _infer_describe_type(col_values)})

        sample = [dict(zip(headers, row)) for row in rows[:3]]

        results.append({
            "sheet": ws.title,
            "columns": columns,
            "row_count": len(rows),
            "sample": sample,
        })

    return results[0] if len(results) == 1 else results


@mcp.tool()
def sql_query(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    sql: Annotated[str, Field(description=(
        "SQL SELECT statement to execute. Each sheet is a table (quote names "
        "with double quotes if they contain spaces, e.g. '\"Q1 Sales\"'). "
        "Supports WHERE, ORDER BY, LIMIT, GROUP BY, HAVING, JOINs across "
        "sheets, DISTINCT, UNION, subqueries, and aggregates (COUNT, SUM, "
        "AVG, MIN, MAX). "
        "Example: SELECT name, revenue FROM Sales WHERE status = 'Active' "
        "ORDER BY revenue DESC LIMIT 20"
    ))],
    header_row: Annotated[int, Field(description="1-based row number containing column headers. Defaults to 1.")] = 1,
) -> list[dict]:
    """Execute a read-only SQL SELECT query against the spreadsheet data.

    Every sheet in the workbook is loaded as a database table, with the
    header row defining column names and data rows below it. Returns
    results as a list of {column: value} objects.

    Only SELECT (and WITH ... SELECT) statements are accepted. Use
    sql_execute for INSERT, UPDATE, or DELETE.
    """
    sql_stripped = sql.strip().rstrip(";")
    first_keyword = sql_stripped.split()[0].upper() if sql_stripped else ""
    if first_keyword not in ("SELECT", "WITH"):
        raise ValueError(
            "sql_query only accepts SELECT statements (or WITH ... SELECT). "
            "Use sql_execute for INSERT/UPDATE/DELETE."
        )

    wb = _load(file)
    conn = _load_sheets_to_duckdb(wb, header_row)

    result = conn.execute(sql_stripped)
    columns = [desc[0] for desc in result.description]
    return [dict(zip(columns, row)) for row in result.fetchall()]


@mcp.tool()
def sql_execute(
    file: Annotated[str, Field(description="Path to the .xlsx file")],
    sql: Annotated[str, Field(description=(
        "SQL mutation statement to execute: INSERT INTO, UPDATE, or "
        "DELETE FROM. Sheet names are table names. "
        "Example: UPDATE Sales SET status = 'Closed' "
        "WHERE quarter = 'Q1' AND revenue < 1000"
    ))],
    header_row: Annotated[int, Field(description="1-based row number containing column headers. Defaults to 1.")] = 1,
) -> dict:
    """Execute a mutating SQL statement and write changes back to the file.

    Supports INSERT INTO (adds rows), UPDATE (modifies cell values), and
    DELETE FROM (removes rows). The target sheet is determined from the
    SQL statement. After execution, the modified table is written back to
    the .xlsx file atomically. Returns {"affected_rows": N}.

    All sheets are available for cross-sheet references (e.g. INSERT INTO
    Summary SELECT ... FROM Transactions).
    """
    sql_stripped = sql.strip().rstrip(";")
    target_table = _extract_target_table(sql_stripped)

    wb = _load(file)
    ws = _resolve_sheet(wb, target_table)

    headers, _ = _sheet_to_records(ws, header_row)
    if not headers:
        raise ValueError(f"Sheet {target_table!r} has no headers at row {header_row}")
    headers = _dedup_headers(headers)
    num_cols = len(headers)

    conn = _load_sheets_to_duckdb(wb, header_row)

    result = conn.execute(sql_stripped)
    affected = result.fetchone()[0]

    col_list = ", ".join(f'"{h}"' for h in headers)
    new_rows = conn.execute(f'SELECT {col_list} FROM "{target_table}"').fetchall()

    old_max_row = ws.max_row or header_row
    for r in range(header_row + 1, old_max_row + 1):
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).value = None

    for r_idx, row in enumerate(new_rows):
        for c_idx, val in enumerate(row):
            ws.cell(row=header_row + 1 + r_idx, column=c_idx + 1, value=val)

    _save(wb, file)
    return {"affected_rows": affected}


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    mcp.run()
