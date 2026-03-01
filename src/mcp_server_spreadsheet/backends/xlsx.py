"""XLSX backend using openpyxl."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook as _OpenpyxlWorkbook
from openpyxl import load_workbook as _openpyxl_load

from .base import SpreadsheetSheet, SpreadsheetWorkbook


class XlsxSheet(SpreadsheetSheet):
    """Wraps an openpyxl Worksheet."""

    def __init__(self, ws) -> None:
        self._ws = ws

    @property
    def title(self) -> str:
        return self._ws.title

    @title.setter
    def title(self, value: str) -> None:
        self._ws.title = value

    @property
    def max_row(self) -> int:
        return self._ws.max_row or 0

    @property
    def max_column(self) -> int:
        return self._ws.max_column or 0

    def cell_value(self, row: int, col: int) -> Any:
        return self._ws.cell(row=row, column=col).value

    def set_cell(self, row: int, col: int, value: Any) -> None:
        self._ws.cell(row=row, column=col, value=value)

    def iter_rows(
        self,
        *,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ) -> Iterator[tuple]:
        yield from self._ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )

    def append(self, values: list) -> None:
        self._ws.append(values)

    def insert_rows(self, row: int, count: int) -> None:
        self._ws.insert_rows(row, amount=count)

    def delete_rows(self, row: int, count: int) -> None:
        self._ws.delete_rows(row, amount=count)

    def insert_cols(self, col: int, count: int) -> None:
        self._ws.insert_cols(col, amount=count)

    def delete_cols(self, col: int, count: int) -> None:
        self._ws.delete_cols(col, amount=count)


class XlsxWorkbook(SpreadsheetWorkbook):
    """Wraps an openpyxl Workbook."""

    def __init__(self, wb: _OpenpyxlWorkbook) -> None:
        self._wb = wb

    @property
    def sheetnames(self) -> list[str]:
        return self._wb.sheetnames

    @property
    def worksheets(self) -> list[SpreadsheetSheet]:
        return [XlsxSheet(ws) for ws in self._wb.worksheets]

    def get_sheet(self, name: str) -> XlsxSheet:
        if name not in self._wb.sheetnames:
            raise ValueError(
                f"Sheet not found: {name!r}. Available: {self._wb.sheetnames}"
            )
        return XlsxSheet(self._wb[name])

    def create_sheet(
        self, title: str | None = None, index: int | None = None,
    ) -> XlsxSheet:
        ws = self._wb.create_sheet(title=title, index=index)
        return XlsxSheet(ws)

    def delete_sheet(self, name: str) -> None:
        if name not in self._wb.sheetnames:
            raise ValueError(f"Sheet not found: {name!r}")
        del self._wb[name]

    def copy_sheet(self, source_name: str) -> XlsxSheet:
        if source_name not in self._wb.sheetnames:
            raise ValueError(f"Sheet not found: {source_name!r}")
        copied = self._wb.copy_worksheet(self._wb[source_name])
        return XlsxSheet(copied)

    def move_sheet(self, sheet: SpreadsheetSheet, offset: int) -> None:
        self._wb.move_sheet(sheet.title, offset=offset)

    def save(self, path: str) -> None:
        p = Path(path)
        fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=p.parent)
        os.close(fd)
        try:
            self._wb.save(tmp)
            os.replace(tmp, p)
        except BaseException:
            os.unlink(tmp)
            raise

    @classmethod
    def load(cls, path: str) -> XlsxWorkbook:
        p = Path(path)
        if not p.exists():
            raise ValueError(f"File not found: {path}")
        return cls(_openpyxl_load(p, data_only=False))

    @classmethod
    def create(cls, sheet_name: str | None = None) -> XlsxWorkbook:
        wb = _OpenpyxlWorkbook()
        if sheet_name:
            wb.active.title = sheet_name
        return cls(wb)
